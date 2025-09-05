# -*- coding: utf-8 -*-
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from flask import Flask, render_template, request, redirect, url_for, flash as _flask_flash, make_response, session, send_file
from flask_wtf.csrf import CSRFProtect, generate_csrf
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from markupsafe import Markup

from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.routing import BuildError
from sqlalchemy import or_, and_, event, text
from uuid import uuid4

app = Flask(__name__)

# --- Security & safety hardening ---
# Core security-related config
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=True if os.environ.get('SESSION_COOKIE_SECURE','0') == '1' else False,
    MAX_CONTENT_LENGTH=int(os.environ.get('MAX_CONTENT_LENGTH', str(10*1024*1024))),  # 10 MB
)

# CSRF protection for all POST/PUT/PATCH/DELETE
csrf = CSRFProtect(app)

# Rate limiting (not applied globally; per-route where needed)
limiter = Limiter(get_remote_address, app=app, default_limits=[])

# Security headers
@app.after_request
def _set_secure_headers(resp):
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("X-Frame-Options", "DENY")
    resp.headers.setdefault("X-XSS-Protection", "1; mode=block")
    # NOTE: CSP kept minimal to avoid breaking inline templates; adjust for prod if needed
    resp.headers.setdefault("Content-Security-Policy", "default-src 'self'; style-src 'self' 'unsafe-inline'; img-src 'self' data:;")
    resp.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    resp.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return resp

# Expose csrf helpers in templates
@app.context_processor
def inject_csrf():
    return {
        'csrf_token': lambda: generate_csrf(),
        'csrf_field': lambda: Markup(f'<input type="hidden" name="csrf_token" value="{generate_csrf()}">'),
    }
VERSION = 'v23'
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///inventory.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"

# Enable SQLite foreign keys
def _enable_sqlite_fk(dbapi_conn, conn_record):
    try:
        cur = dbapi_conn.cursor()
        cur.execute("PRAGMA foreign_keys=ON")
        cur.close()
    except Exception:
        pass

try:
    event.listen(db.engine, "connect", _enable_sqlite_fk)
except Exception:
    pass

# Safe url_for for templates
def safe_url_for(endpoint, **values):
    try:
        return url_for(endpoint, **values)
    except BuildError:
        return None

app.jinja_env.globals.update(safe_url_for=safe_url_for, VERSION=VERSION)

# --- Timezone helpers: store UTC (naive) but display in Europe/Moscow
def _to_msk_str(dt, fmt='%d.%m.%Y %H:%M:%S'):
    """Return formatted string in Europe/Moscow for a datetime `dt`.
    If `dt` is naive it's assumed to be UTC (this project stores UTC naive datetimes).
    """
    if not dt:
        return ''
    try:
        # ensure tz-aware UTC
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo('UTC'))
        msk = dt.astimezone(ZoneInfo('Europe/Moscow'))
        return msk.strftime(fmt)
    except Exception:
        try:
            return dt.strftime(fmt)
        except Exception:
            return ''

# Jinja filter: usage in templates: {{ some_dt|msk("%d.%m.%Y %H:%M") }}
def _jinja_msk_filter(dt, fmt='%d.%m.%Y %H:%M:%S'):
    return _to_msk_str(dt, fmt)

app.jinja_env.filters['msk'] = _jinja_msk_filter

import re as _re
import re

# Flash message normalizer (clean)
_sale_re = re.compile(r'.*:\\s*(.+)\\sx(\\d+)\\.\\s.*:\\s*([0-9]+(?:\\.[0-9]+)?)\\s*$')
_inc_re = re.compile(r'.*\\s\\+(\\d+)\\.$')
_dec_re = re.compile(r'.*\\s\\-(\\d+)\\.$')

def _fix_message(msg):
    try:
        s = str(msg)
    except Exception:
        return msg
    # Exact known garbled -> clean Russian
    exact = {
        "??? ????? ? ?? ?????.": "Недостаточно прав для доступа.",
        "Вход выполнен успешно.": "Вход выполнен успешно.",
        "Неверное имя пользователя или пароль.": "Неверное имя пользователя или пароль.",
        "Вы вышли из системы.": "Вы вышли из системы.",
        "Имя пользователя и пароль обязательны.": "Название обязательно.",
        "????? ????????.": "Товар добавлен.",
        "????? ???????.": "Товар обновлён.",
        "????? ???????.": "Товар архивирован.",
        "????? ?????????? ?? ????.": "Товар разархивирован.",
        "????? ?????? ???? 0.": "Нельзя уменьшить ниже 0.",
        "? ??? ???? ????? - ?? ??????? (????? ??????).": "Товар имеет продажи — он будет архивирован (вместо удаления).",
        "Пользователь удалён.": "Товар удалён.",
        "?????? ??? 'from' (YYYY-MM-DD).": "Некорректная дата 'from' (YYYY-MM-DD).",
        "?????? ??? 'to' (YYYY-MM-DD).": "Некорректная дата 'to' (YYYY-MM-DD).",
        "????? ? ???? ? ?? ????? ???? ????.": "Товар в архиве и недоступен для продажи.",
        "??????? ?????? ???? Пользователь удалён.": "Количество должно быть больше нуля.",
        "Должен остаться хотя бы один администратор.": "Недостаточно товара на складе.",
        "?????? ? % ?????? ???? ?? 0 ?? 100.": "Скидка в % должна быть от 0 до 100.",
        "?????? ? ??? ?????? ???? ?????????.": "Скидка в сумме должна быть неотрицательной.",
        "?Пользователь с таким именем уже существует.": "Итоговая цена не может быть отрицательной.",
        "Нельзя удалить собственную учетную запись.": "Файл не выбран.",

    }
    if s in exact:
        return exact[s]
    # Common cases
    if 'xlsx, .xls или .csv' in s:
        return 'Неподдерживаемый формат файла. Используйте .xlsx, .xls или .csv'
    if "'from'" in s and 'YYYY' in s:
        return 'Некорректная дата \'from\' (YYYY-MM-DD).'
    if "'to'" in s and 'YYYY' in s:
        return 'Некорректная дата \'to\' (YYYY-MM-DD).'
    m = _sale_re.match(s)
    if m:
        name, qty, total = m.groups()
        return f'Продажа оформлена: {name} x{qty}. Сумма: {total}'
    m = _inc_re.match(s)
    if m:
        n = m.group(1)
        return f'Остаток увеличен на +{n}.'
    m = _dec_re.match(s)
    if m:
        n = m.group(1)
        return f'Остаток уменьшен на -{n}.'
    return s

def _fix_message2(msg):
    try:
        s = str(msg)
    except Exception:
        return msg
    exact = {
        "Недостаточно прав для доступа.": "Недостаточно прав для доступа.",
        "Вход выполнен успешно.": "Вход выполнен успешно.",
        "Неверное имя пользователя или пароль.": "Неверное имя пользователя или пароль.",
        "Вы вышли из системы.": "Вы вышли из системы.",
        "Название обязательно.": "Название обязательно.",
        "Товар добавлен.": "Товар добавлен.",
        "Товар обновлён.": "Товар обновлён.",
        "Товар архивирован.": "Товар архивирован.",
        "Товар разархивирован.": "Товар разархивирован.",
        "Нельзя уменьшить ниже 0.": "Нельзя уменьшить ниже 0.",
        "Товар имеет продажи — он будет архивирован (вместо удаления).": "Товар имеет продажи — он будет архивирован (вместо удаления).",
        "Товар удалён.": "Товар удалён.",
        "Некорректная дата 'from' (YYYY-MM-DD).": "Некорректная дата 'from' (YYYY-MM-DD).",
        "Некорректная дата 'to' (YYYY-MM-DD).": "Некорректная дата 'to' (YYYY-MM-DD).",
        "Товар в архиве и недоступен для продажи.": "Товар в архиве и недоступен для продажи.",
        "Количество должно быть больше нуля.": "Количество должно быть больше нуля.",
        "Недостаточно товара на складе.": "Недостаточно товара на складе.",
        "Скидка в % должна быть от 0 до 100.": "Скидка в % должна быть от 0 до 100.",
        "Скидка в сумме должна быть неотрицательной.": "Скидка в сумме должна быть неотрицательной.",
        "Итоговая цена не может быть отрицательной.": "Итоговая цена не может быть отрицательной.",
        "Файл не выбран.": "Файл не выбран.",

    }
    if s in exact:
        return exact[s]
    if 'xlsx, .xls или .csv' in s:
        return 'Неподдерживаемый формат файла. Используйте .xlsx, .xls или .csv'
    if "'from'" in s and 'YYYY' in s:
        return "Некорректная дата 'from' (YYYY-MM-DD)."
    if "'to'" in s and 'YYYY' in s:
        return "Некорректная дата 'to' (YYYY-MM-DD)."
    m = _sale_re.match(s)
    if m:
        name, qty, total = m.groups()
        return f"Продажа оформлена: {name} x{qty}. Сумма: {total}"
    m = _inc_re.match(s)
    if m:
        n = m.group(1)
        return f"Остаток увеличен на +{n}."
    m = _dec_re.match(s)
    if m:
        n = m.group(1)
        return f"Остаток уменьшен на -{n}."
    return s

def flash(message, category='message'):
    try:
        s = str(message)
    except Exception:
        s = message
    return _flask_flash(s, category)
def _normalize_text(s: str) -> str:
    if not s: return ''
    s = s.lower()
    # remove spaces and common separators
    s = _re.sub(r'[\s\-\._/\\,;:]+', '', s)
    return s


# ----------------- Models -----------------
class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='cashier')
    def set_password(self, pwd): self.password_hash = generate_password_hash(pwd)
    def check_password(self, pwd): return check_password_hash(self.password_hash, pwd)

class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False, index=True)
    sku = db.Column(db.String(120), index=True)
    barcode = db.Column(db.String(120), index=True)
    image_url = db.Column(db.String(500))
    search_text = db.Column(db.Text, nullable=False, default='')
    cost_price = db.Column(db.Float, nullable=False, default=0.0)
    base_price = db.Column(db.Float, nullable=False, default=0.0)
    stock_qty = db.Column(db.Integer, nullable=False, default=0)
    is_archived = db.Column(db.Boolean, nullable=False, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

class Sale(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False, default=1)
    unit_price_sold = db.Column(db.Float, nullable=False)
    discount_type = db.Column(db.String(10))
    discount_value = db.Column(db.Float)
    final_unit_price = db.Column(db.Float, nullable=False)
    total_price = db.Column(db.Float, nullable=False)
    profit = db.Column(db.Float, nullable=False)
    cashier = db.Column(db.String(80))
    note = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    product = db.relationship('Product', backref='sales')

class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user = db.Column(db.String(80))
    action = db.Column(db.String(80), nullable=False)
    entity = db.Column(db.String(80))
    entity_id = db.Column(db.Integer)
    description = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

class RentalItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False, unique=True)
    code = db.Column(db.String(50), index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

class Rental(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('rental_item.id'), nullable=False)
    renter_name = db.Column(db.String(200))
    renter_phone = db.Column(db.String(50))
    started_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    returned_at = db.Column(db.DateTime)
    total_price = db.Column(db.Float, default=0.0, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    comment = db.Column(db.Text)
    item = db.relationship('RentalItem', backref='rentals')

class RentalCharge(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    rental_id = db.Column(db.Integer, db.ForeignKey('rental.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    label = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    rental = db.relationship('Rental', backref='charges')

@login_manager.user_loader
def load_user(uid):
    return db.session.get(User, int(uid))

# --------------- Helpers -----------------
def roles_required(*roles):
    def deco(fn):
        from functools import wraps
        @wraps(fn)
        def inner(*a, **kw):
            if not current_user.is_authenticated:
                return login_manager.unauthorized()
            if current_user.role not in roles:
                flash("Недостаточно прав для доступа.", "error")
                return redirect(url_for('index'))
            return fn(*a, **kw)
        return inner
    return deco

def log_action(action, entity=None, entity_id=None, description=None):
    try:
        u = current_user.username if current_user.is_authenticated else None
    except Exception:
        u = None
    db.session.add(AuditLog(user=u, action=action, entity=entity, entity_id=entity_id, description=description))
    db.session.commit()

def apply_sqlite_migrations():
    db.create_all()
    with db.engine.connect() as conn:
        try:
            cols = [row[1] for row in conn.execute(text("PRAGMA table_info(product)"))]
            if 'is_archived' not in cols:
                conn.execute(text("ALTER TABLE product ADD COLUMN is_archived BOOLEAN NOT NULL DEFAULT 0"))
            if 'barcode' not in cols:
                conn.execute(text("ALTER TABLE product ADD COLUMN barcode VARCHAR(120)"))
            if 'image_url' not in cols:
                conn.execute(text("ALTER TABLE product ADD COLUMN image_url VARCHAR(500)"))
            if 'search_text' not in cols:
                conn.execute(text("ALTER TABLE product ADD COLUMN search_text TEXT NOT NULL DEFAULT ''"))
            # Sale table migrations
            cols_s = [row[1] for row in conn.execute(text("PRAGMA table_info(sale)"))]
            if 'note' not in cols_s:
                conn.execute(text("ALTER TABLE sale ADD COLUMN note TEXT"))
            # Rental table migrations
            cols_r = [row[1] for row in conn.execute(text("PRAGMA table_info(rental)"))]
            if 'comment' not in cols_r:
                conn.execute(text("ALTER TABLE rental ADD COLUMN comment TEXT"))
            if 'renter_name' not in cols_r:
                conn.execute(text("ALTER TABLE rental ADD COLUMN renter_name VARCHAR(200)"))
        except Exception:
            pass

def init_defaults():
    db.create_all()
    apply_sqlite_migrations()
    if not User.query.filter_by(username='admin').first():
        u = User(username='admin', role='admin'); u.set_password(os.environ.get('ADMIN_PASSWORD','admin123')); db.session.add(u)
    if not User.query.filter_by(username='cashier').first():
        u = User(username='cashier', role='cashier'); u.set_password(os.environ.get('CASHIER_PASSWORD','cashier123')); db.session.add(u)
    db.session.commit()
    # Seed default rental items
    try:
        defaults = [
            ('PS5', 'PS5'),
            ('PS5M', 'PS5M'),
            ('PS5P', 'PS5P'),
            ('PS5Y', 'PS5Y'),
            ('PS5Большая', 'PS5BIG'),
            ('Xbosx seres S', 'XBSS'),
            ('Проектор', 'PROJECTOR'),
        ]
        for name, code in defaults:
            if not db.session.query(RentalItem).filter_by(name=name).first():
                db.session.add(RentalItem(name=name, code=code))
        db.session.commit()
    except Exception:
        db.session.rollback()


# ---- Import helpers (preview/apply) ----
def _ensure_import_tmp():
    tmp = os.path.join(app.instance_path, 'import_tmp')
    try:
        os.makedirs(tmp, exist_ok=True)
    except Exception:
        pass
    return tmp

def _parse_product_rows_from_filelike(file, ext: str):
    """
    Returns: list of dict rows with keys:
      name, sku, barcode, image_url, cost_price, base_price, stock_qty
    Does not commit to DB. Raises on fatal parse errors.
    """
    def norm_header(s):
        try:
            s = '' if s is None else str(s)
        except Exception:
            s = ''
        return s.strip().lower()

    def to_float(x):
        try:
            return float(str(x).replace(',', '.'))
        except Exception:
            return None

    def to_int(x):
        try:
            return int(float(str(x).replace(',', '.')))
        except Exception:
            return None

    rows = []
    if ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        values_rows = list(ws.iter_rows(values_only=True))
        header_row = 0
        headers = []
        best = -1
        for ridx, row in enumerate(values_rows[:50]):
            heads_try = [norm_header(v) for v in row]
            score = sum(1 for h in heads_try if h)
            if score > best:
                best = score
                header_row = ridx
                headers = heads_try
        synonyms = {
            'name': ['name','товар','наименование','название','product'],
            'sku': ['sku','код','артикул','product_code','код товара'],
            'barcode': ['barcode','штрихкод','штрих-код','штрих код','ean','upc'],
            'cost_price': ['cost_price','cost','себестоимость','закуп','закупка','закупочная','покупка','закуп.'],
            'base_price': ['base_price','price','цена','цена продажи','розничная цена','розница','продажа'],
            'stock_qty': ['stock_qty','qty','количество','кол-во','остаток','доступно','остаток на складе','кол-во на складе'],
            'image_url': ['image_url','image','photo','фото','изображение','url']
        }
        idx = {}
        for k, alts in synonyms.items():
            for a in alts:
                if a in headers:
                    idx[k] = headers.index(a); break
        # fallback auto-map ru
        for i, h in enumerate(headers):
            if any(x in h for x in ('наимен','товар','назван')): idx.setdefault('name', i)
            if ('код' in h and 'штрих' not in h) or 'артикул' in h: idx.setdefault('sku', i)
            if 'штрих' in h: idx.setdefault('barcode', i)
            if any(x in h for x in ('себесто','закуп','покуп')): idx.setdefault('cost_price', i)
            if ('цена' in h and 'сумм' not in h) or 'розниц' in h or 'продаж' in h: idx.setdefault('base_price', i)
            if any(x in h for x in ('остаток','доступ','кол-во','количество')): idx.setdefault('stock_qty', i)
        for ridx, row in enumerate(values_rows[header_row+1:], start=header_row+1):
            get = lambda key: row[idx[key]] if key in idx and idx[key] < len(row) else None
            name = (get('name') or '').strip() if get('name') else ''
            if not name: continue
            sku = (get('sku') or None) or None
            barcode = (get('barcode') or None) or None
            image_url = (get('image_url') or None) or None
            cost = to_float(get('cost_price'))
            base = to_float(get('base_price'))
            qty = to_int(get('stock_qty'))
            rows.append(dict(name=name, sku=sku, barcode=barcode, image_url=image_url,
                             cost_price=cost, base_price=base, stock_qty=qty))
    elif ext == '.xls':
        try:
            import xlrd
        except Exception as e:
            raise RuntimeError('Для импорта .xls установите пакет xlrd==1.2.0') from e
        content = file.read()
        wb = xlrd.open_workbook(file_contents=content)
        sh = wb.sheet_by_index(0)
        headers = [norm_header(sh.cell_value(0, c)) for c in range(sh.ncols)]
        synonyms = {
            'name': ['name','товар','наименование','название','product'],
            'sku': ['sku','код','артикул','product_code','код товара'],
            'barcode': ['barcode','штрихкод','штрих-код','штрих код','ean','upc'],
            'cost_price': ['cost_price','cost','себестоимость','закуп','закупка','закупочная','покупка','закуп.'],
            'base_price': ['base_price','price','цена','цена продажи','розничная цена','розница','продажа'],
            'stock_qty': ['stock_qty','qty','количество','кол-во','остаток','доступно','остаток на складе','кол-во на складе'],
            'image_url': ['image_url','image','photo','фото','изображение','url']
        }
        idx = {}
        for k, alts in synonyms.items():
            for a in alts:
                if a in headers:
                    idx[k] = headers.index(a); break
        for r in range(1, sh.nrows):
            get = lambda key: sh.cell_value(r, idx[key]) if key in idx else None
            name = (get('name') or '').strip()
            if not name: continue
            sku = (get('sku') or '').strip() or None
            barcode = (get('barcode') or '').strip() or None
            image_url = (get('image_url') or '').strip() or None
            cost = to_float(get('cost_price'))
            base = to_float(get('base_price'))
            qty = to_int(get('stock_qty'))
            rows.append(dict(name=name, sku=sku, barcode=barcode, image_url=image_url,
                             cost_price=cost, base_price=base, stock_qty=qty))
    elif ext == '.csv':
        import csv, io
        content = file.read()
        try:
            s = content.decode('utf-8-sig')
        except Exception:
            s = content.decode('cp1251')
        reader = csv.DictReader(io.StringIO(s))
        def lc(d): return {norm_header(k): v for k,v in d.items()}
        for row in reader:
            rr = lc(row)
            csv_map = {
                'name': ['name','товар','наименование','название','product'],
                'sku': ['sku','код','артикул','product_code','код товара'],
                'barcode': ['barcode','штрихкод','штрих-код','штрих код','ean','upc'],
                'cost_price': ['cost_price','cost','себестоимость','закуп','закупка','закупочная','покупка','закуп.'],
                'base_price': ['base_price','price','цена','цена продажи','розничная цена','розница','продажа'],
                'stock_qty': ['stock_qty','qty','количество','кол-во','остаток','доступно','остаток на складе','кол-во на складе'],
                'image_url': ['image_url','image','photo','фото','изображение','url']
            }
            for key, alts in csv_map.items():
                if rr.get(key) not in (None, ''):
                    continue
                for a in alts:
                    if rr.get(a) not in (None, ''):
                        rr[key] = rr.get(a); break
            name = (rr.get('name') or rr.get('название') or rr.get('наименование') or '').strip()
            if not name: continue
            sku = (rr.get('sku') or rr.get('артикул') or '').strip() or None
            barcode = (rr.get('barcode') or rr.get('штрихкод') or rr.get('штрих-код') or '').strip() or None
            image_url = (rr.get('image_url') or rr.get('image') or rr.get('фото') or '').strip() or None
            cost = to_float(rr.get('cost_price'))
            base = to_float(rr.get('base_price'))
            qty = None
            try:
                qty = int(float((rr.get('stock_qty') or '').replace(',', '.')))
            except Exception:
                qty = None
            rows.append(dict(name=name, sku=sku, barcode=barcode, image_url=image_url,
                             cost_price=cost, base_price=base, stock_qty=qty))
    else:
        raise RuntimeError('Неподдерживаемый формат файла. Используйте .xlsx, .xls или .csv')
    return rows
# --------------- Routes -----------------
@app.route('/version')
def app_version():
    return {'version': VERSION}, 200

@app.route('/health')
def health():
    return {'status': 'ok', 'version': VERSION}, 200


def backfill_search_text():
    try:
        for p in Product.query.all():
            need = _normalize_text(' '.join(filter(None, [p.name, p.sku, p.barcode])))
            if p.search_text != need:
                p.search_text = need
        db.session.commit()
    except Exception:
        pass


@app.route('/')
@login_required
def index():
    return render_template('index.html')

# ---- Auth ----
@app.route('/login', methods=['GET','POST'])
@limiter.limit('5 per minute')
def login():
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            flash('Вход выполнен успешно.', 'success')
            return redirect(url_for('index'))
        flash('Неверное имя пользователя или пароль.', 'error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Вы вышли из системы.', 'success')
    return redirect(url_for('login'))

# ---- Admin: Products ----
@app.route('/admin/products', methods=['GET'])
@login_required
@roles_required('admin')
def admin_products():
    q = request.args.get('q','').strip()
    show_archived = request.args.get('archived','0') == '1'
    query = Product.query
    if not show_archived:
        query = query.filter(Product.is_archived.is_(False))
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Product.name.ilike(like), Product.sku.ilike(like)))
    # Smart ordering: in-stock first, then by stock qty desc, then name
    from sqlalchemy import case
    stock_rank = case((Product.stock_qty > 0, 0), else_=1)
    products = query.order_by(stock_rank.asc(), Product.stock_qty.desc(), Product.name.asc()).all()
    return render_template('admin_products.html', products=products, q=q, show_archived=show_archived)

@app.route('/admin/products/add', methods=['POST'])
@login_required
@roles_required('admin')
def admin_products_add():
    name = request.form.get('name','').strip()
    sku = request.form.get('sku','').strip() or None
    barcode = request.form.get('barcode','').strip() or None
    image_url = request.form.get('image_url','').strip() or None
    cost_price = float((request.form.get('cost_price') or '0').replace(',', '.'))
    base_price = float((request.form.get('base_price') or '0').replace(',', '.'))
    stock_qty = int(request.form.get('stock_qty') or 0)
    if not name:
        flash("Название обязательно.", "error")
        return redirect(url_for('admin_products'))
    p = Product(name=name, sku=sku, barcode=barcode, image_url=image_url, cost_price=cost_price, base_price=base_price, stock_qty=stock_qty)
    p.search_text = _normalize_text(' '.join(filter(None, [p.name, p.sku, p.barcode])))
    db.session.add(p); db.session.commit()
    # For single add, define counters to satisfy shared logging/flash format
    added, updated, skipped = 1, 0, 0
    log_action('product_import', 'Product', None, f'added={added}, updated={updated}, skipped={skipped}')
    flash(f'Импорт завершён: добавлено {added}, обновлено {updated}, пропущено {skipped}.', 'success')
    return redirect(url_for('admin_products'))

@app.route('/admin/products/<int:pid>/edit', methods=['GET','POST'])
@login_required
@roles_required('admin')
def admin_product_edit(pid):
    p = Product.query.get_or_404(pid)
    if request.method == 'POST':
        name = request.form.get('name','').strip()
        sku = request.form.get('sku','').strip() or None
        barcode = request.form.get('barcode','').strip() or None
        image_url = request.form.get('image_url','').strip() or None
        cost_price = float(request.form.get('cost_price') or p.cost_price)
        base_price = float(request.form.get('base_price') or p.base_price)
        stock_qty = int(request.form.get('stock_qty') or p.stock_qty)
        if not name:
            flash("Название обязательно.", "error")
            return redirect(url_for('admin_product_edit', pid=pid))
        changes = []
        if p.cost_price != cost_price: changes.append(f"cost_price: {p.cost_price}->{cost_price}")
        if p.base_price != base_price: changes.append(f"base_price: {p.base_price}->{base_price}")
        if p.barcode != barcode: changes.append(f"barcode: {p.barcode}->{barcode}")
        p.name, p.sku, p.barcode, p.image_url, p.cost_price, p.base_price, p.stock_qty = name, sku, barcode, image_url, cost_price, base_price, stock_qty
        p.search_text = _normalize_text(' '.join(filter(None, [p.name, p.sku, p.barcode])))
        db.session.commit()
        log_action('product_update', 'Product', p.id, "; ".join(changes) or "updated")
        flash("Товар обновлён.", "success")
        return redirect(url_for('admin_products'))
    return render_template('product_edit.html', p=p)

@app.route('/admin/products/<int:pid>/archive', methods=['POST'])
@login_required
@roles_required('admin')
def admin_products_archive(pid):
    p = db.session.get(Product, pid)
    if not p:
        flash('Товар не найден.', 'error')
        return redirect(url_for('admin_products'))
    
    try:
        print(f"Архивирую товар: ID={p.id}, Name={p.name}, Current archived status={p.is_archived}")
        p.is_archived = True
        db.session.commit()
        print(f"Новый статус архивации: {p.is_archived}")
        log_action('product_archive', 'Product', p.id, f'Archived {p.name}')
        flash("Товар архивирован.", "success")
    except Exception as e:
        db.session.rollback()
        print(f"Ошибка при архивации: {e}")
        flash(f'Ошибка при архивации товара: {e}', 'error')
    
    return redirect(url_for('admin_products'))

@app.route('/admin/products/<int:pid>/unarchive', methods=['POST'])
@login_required
@roles_required('admin')
def admin_products_unarchive(pid):
    p = db.session.get(Product, pid)
    if not p:
        flash('Товар не найден.', 'error')
        return redirect(url_for('admin_products'))
    
    try:
        print(f"Разархивирую товар: ID={p.id}, Name={p.name}, Current archived status={p.is_archived}")
        p.is_archived = False
        db.session.commit()
        print(f"Новый статус архивации: {p.is_archived}")
        log_action('product_unarchive', 'Product', p.id, f'Unarchived {p.name}')
        flash("Товар разархивирован.", "success")
    except Exception as e:
        db.session.rollback()
        print(f"Ошибка при разархивации: {e}")
        flash(f'Ошибка при разархивации товара: {e}', 'error')
    
    return redirect(url_for('admin_products'))


@app.route('/admin/products/<int:pid>/stock/inc', methods=['POST'])
@login_required
@roles_required('admin')
def admin_products_stock_inc(pid):
    p = Product.query.get_or_404(pid)
    step = int(request.form.get('step', 1) or 1)
    if step < 1: step = 1
    p.stock_qty = (p.stock_qty or 0) + step
    db.session.commit()
    log_action('stock_inc','Product',p.id,f'+{step}, new stock={p.stock_qty}')
    flash(f"Остаток увеличен на +{step}.", "success")
    return redirect(url_for('admin_products', archived=request.args.get('archived','0')))

@app.route('/admin/products/<int:pid>/stock/dec', methods=['POST'])
@login_required
@roles_required('admin')
def admin_products_stock_dec(pid):
    p = Product.query.get_or_404(pid)
    step = int(request.form.get('step', 1) or 1)
    if step < 1: step = 1
    new_qty = (p.stock_qty or 0) - step
    if new_qty < 0:
        flash("Нельзя уменьшить ниже 0.", "error")
        return redirect(url_for('admin_products', archived=request.args.get('archived','0')))
    p.stock_qty = new_qty
    db.session.commit()
    log_action('stock_dec','Product',p.id,f'-{step}, new stock={p.stock_qty}')
    flash(f"Остаток уменьшен на -{step}.", "success")
    return redirect(url_for('admin_products', archived=request.args.get('archived','0')))

@app.route('/admin/products/<int:pid>/delete', methods=['POST'])
@login_required
@roles_required('admin')

def admin_products_delete(pid):
    p = Product.query.get_or_404(pid)
    # Unified policy: never hard-delete products; archive instead to keep history consistent
    if p.is_archived:
        flash("Товар уже в архиве.", "info")
        return redirect(url_for('admin_products', archived=request.args.get('archived','0')))
    p.is_archived = True
    db.session.commit()
    log_action('product_archive_on_delete','Product',p.id,'Archived via delete')
    flash("Товар архивирован (вместо удаления).", "success")
    return redirect(url_for('admin_products', archived=request.args.get('archived','0')))
    name = p.name
    db.session.delete(p); db.session.commit()
    log_action('product_delete','Product',pid,f'Deleted {name}')
    flash("Товар удалён.", "success")
    return redirect(url_for('admin_products'))

# ---- Reports & Activity ----
@app.route('/admin/reports')
@login_required
@roles_required('admin')
def admin_reports():
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    q = Sale.query
    if date_from:
        try:
            q = q.filter(Sale.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            flash("Некорректная дата 'from' (YYYY-MM-DD).", "error")
    if date_to:
        try:
            dt = datetime.fromisoformat(date_to)
            q = q.filter(Sale.created_at < dt.replace(hour=23, minute=59, second=59))
        except ValueError:
            flash("Некорректная дата 'to' (YYYY-MM-DD).", "error")
    sales = q.order_by(Sale.created_at.desc()).all()
    revenue = sum(s.total_price for s in sales)
    profit = sum(s.profit for s in sales)
    return render_template('reports.html', sales=sales, revenue=revenue, profit=profit, date_from=date_from, date_to=date_to)

@app.route('/admin/reports/export.csv')
@login_required
@roles_required('admin')
def admin_reports_export_csv():
    import csv, io
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    q = Sale.query
    if date_from:
        try:
            q = q.filter(Sale.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            flash("Некорректная дата 'from' (YYYY-MM-DD).", "error")
    if date_to:
        try:
            dt = datetime.fromisoformat(date_to)
            q = q.filter(Sale.created_at < dt.replace(hour=23, minute=59, second=59))
        except ValueError:
            flash("Некорректная дата 'to' (YYYY-MM-DD).", "error")
    sales = q.order_by(Sale.created_at.desc()).all()
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Дата/время","Товар","SKU","Кол-во","Цена за ед. (продажа)","Тип скидки","Значение скидки","Итог. цена за ед.","Сумма","Прибыль","Кассир","Примечание"])
    for s in sales:
        try:
            name = s.product.name if s.product else ""
        except Exception:
            name = ""
        note = (s.note or "").replace("\n"," ").strip()
        w.writerow([
            _to_msk_str(s.created_at, '%Y-%m-%d %H:%M:%S') if getattr(s, 'created_at', None) else "",
            name,
            getattr(s.product, 'sku', "") if getattr(s, 'product', None) else "",
            s.quantity,
            s.unit_price_sold,
            s.discount_type or "",
            s.discount_value if s.discount_value is not None else "",
            s.final_unit_price,
            s.total_price,
            s.profit,
            s.cashier or "",
            note
        ])
    data = out.getvalue()
    out.close()
    resp = make_response(data.encode('utf-8-sig'))
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    fname = "sales"
    if date_from or date_to:
        fname += f"_{(date_from or '').replace('-','')}_{(date_to or '').replace('-','')}"
    resp.headers['Content-Disposition'] = f'attachment; filename="{fname}.csv"'
    return resp

@app.route('/admin/activity')
@login_required
@roles_required('admin')
def admin_activity():
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    q = AuditLog.query
    if date_from:
        try:
            q = q.filter(AuditLog.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            flash("Некорректная дата 'from' (YYYY-MM-DD).", "error")
    if date_to:
        try:
            dt = datetime.fromisoformat(date_to)
            q = q.filter(AuditLog.created_at < dt.replace(hour=23, minute=59, second=59))
        except ValueError:
            flash("Некорректная дата 'to' (YYYY-MM-DD).", "error")
    if not date_from and not date_to:
        q = q.order_by(AuditLog.created_at.desc()).limit(200)
        logs = q.all()
    else:
        logs = q.order_by(AuditLog.created_at.desc()).all()
    return render_template('activity.html', logs=logs, date_from=date_from, date_to=date_to)
# ---- Rental ----
@app.route('/admin/activity/export.csv')
@login_required
@roles_required('admin')
def admin_activity_export_csv():
    import csv, io
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    q = AuditLog.query
    if date_from:
        try:
            q = q.filter(AuditLog.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            flash("Некорректная дата 'from' (YYYY-MM-DD).", "error")
    if date_to:
        try:
            dt = datetime.fromisoformat(date_to)
            q = q.filter(AuditLog.created_at < dt.replace(hour=23, minute=59, second=59))
        except ValueError:
            flash("Некорректная дата 'to' (YYYY-MM-DD).", "error")
    logs = q.order_by(AuditLog.created_at.desc()).all()
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Дата/время","Пользователь","Действие","Сущность","ID","Описание"])
    for a in logs:
        w.writerow([
            _to_msk_str(a.created_at, '%Y-%m-%d %H:%M:%S') if getattr(a, 'created_at', None) else "",
            a.user or "",
            a.action or "",
            a.entity or "",
            a.entity_id or "",
            (a.description or "").replace("\n"," ").strip()
        ])
    data = out.getvalue()
    out.close()
    resp = make_response(data.encode('utf-8-sig'))
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    fname = "activity"
    if date_from or date_to:
        fname += f"_{(date_from or '').replace('-','')}_{(date_to or '').replace('-','')}"
    resp.headers['Content-Disposition'] = f'attachment; filename="{fname}.csv"'
    return resp

@app.route('/rental', methods=['GET'])
@login_required
@roles_required('admin','cashier')
def rental():
    items = RentalItem.query.order_by(RentalItem.created_at.asc()).all()
    active = Rental.query.filter(Rental.returned_at.is_(None)).order_by(Rental.started_at.asc()).all()
    recent_closed = Rental.query.filter(Rental.returned_at.is_not(None)).order_by(Rental.returned_at.desc()).limit(20).all()
    return render_template('rental.html', items=items, active=active, recent_closed=recent_closed)

def _parse_dt_local(s):
    try:
        # expected format: YYYY-MM-DDTHH:MM
        # The browser input (datetime-local) gives local time (assume Europe/Moscow here).
        # Convert that local MSK time to UTC naive datetime for storage (project stores UTC naives).
        dt_local = datetime.strptime(s.strip(), '%Y-%m-%dT%H:%M')
        # make tz-aware MSK and convert to UTC
        dt_local = dt_local.replace(tzinfo=ZoneInfo('Europe/Moscow'))
        dt_utc = dt_local.astimezone(ZoneInfo('UTC'))
        # return naive UTC (consistent with existing code which uses datetime.utcnow())
        return dt_utc.replace(tzinfo=None)
    except Exception:
        return None

def _to_float(s):
    try:
        return float(str(s).replace(',', '.'))
    except Exception:
        return None

@app.route('/rental/start', methods=['POST'])
@login_required
@roles_required('admin','cashier')
def rental_start():
    item_id = request.form.get('item_id')
    phone = request.form.get('phone','').strip() or None
    renter_name = (request.form.get('renter_name') or '').strip() or None
    comment = (request.form.get('comment') or '').strip() or None
    amount = _to_float(request.form.get('amount'))
    started_at_str = request.form.get('started_at')
    started_at = _parse_dt_local(started_at_str) if started_at_str else datetime.utcnow()
    if not item_id or amount is None or amount < 0:
        flash('Заполните позицию и сумму.', 'error')
        return redirect(url_for('rental'))
    item = db.session.get(RentalItem, int(item_id))
    if not item:
        flash('Позиция аренды не найдена.', 'error')
        return redirect(url_for('rental'))
    exists = Rental.query.filter_by(item_id=item.id, returned_at=None).first()
    if exists:
        flash('Эта позиция уже в аренде. Сначала верните ее.', 'error')
        return redirect(url_for('rental'))
    r = Rental(item_id=item.id, renter_phone=phone, renter_name=renter_name, started_at=started_at, total_price=0.0, comment=comment)
    db.session.add(r)
    db.session.flush()  # to get r.id
    db.session.add(RentalCharge(rental_id=r.id, amount=amount, label='start'))
    r.total_price = amount
    db.session.commit()
    log_action('rental_start', 'Rental', r.id, f'{item.name} name={renter_name} phone={phone} amount={amount}')
    flash('Аренда оформлена.', 'success')
    return redirect(url_for('rental'))

@app.route('/rental/extend/<int:rid>', methods=['POST'])
@login_required
@roles_required('admin','cashier')
def rental_extend(rid):
    r = db.session.get(Rental, rid)
    if not r or r.returned_at is not None:
        flash('Аренда не найдена или уже закрыта.', 'error')
        return redirect(url_for('rental'))
    amount = _to_float(request.form.get('amount'))
    if amount is None or amount <= 0:
        flash('Укажите сумму продления.', 'error')
        return redirect(url_for('rental'))
    db.session.add(RentalCharge(rental_id=r.id, amount=amount, label='extend'))
    r.total_price = (r.total_price or 0) + amount
    db.session.commit()
    log_action('rental_extend', 'Rental', r.id, f'+{amount}')
    flash('Аренда продлена.', 'success')
    return redirect(url_for('rental'))

@app.route('/rental/<int:rid>/comment', methods=['POST'])
@login_required
def rental_comment(rid):
    print(f"Получен POST запрос для комментария. rid={rid}")
    print(f"Форма: {request.form}")
    
    r = db.session.get(Rental, rid)
    if not r:
        flash('Аренда не найдена.', 'error')
        return redirect(url_for('rental'))
    
    new_comment = (request.form.get('comment') or '').strip() or None
    print(f"Старый комментарий: {r.comment}")
    print(f"Новый комментарий: {new_comment}")
    
    try:
        r.comment = new_comment
        db.session.commit()
        log_action('rental_comment_update', 'Rental', r.id, f'len={len(new_comment or "")}')
        flash('Комментарий сохранён.', 'success')
        print("Комментарий успешно сохранен")
    except Exception as e:
        db.session.rollback()
        flash(f'Не удалось сохранить комментарий: {e}', 'error')
        print(f"Ошибка при сохранении комментария: {e}")
    return redirect(url_for('rental'))

@app.route('/rental/return/<int:rid>', methods=['POST'])
@login_required
@roles_required('admin','cashier')
def rental_return(rid):
    r = db.session.get(Rental, rid)
    if not r or r.returned_at is not None:
        flash('Аренда не найдена или уже закрыта.', 'error')
        return redirect(url_for('rental'))
    r.returned_at = datetime.utcnow()
    db.session.commit()
    log_action('rental_return', 'Rental', r.id, f'total={r.total_price}')
    flash('Аренда закрыта.', 'success')
    return redirect(url_for('rental'))

@app.route('/rental/delete/<int:rid>', methods=['POST'])
@login_required
@roles_required('admin')
def rental_delete(rid):
    r = db.session.get(Rental, rid)
    if not r:
        flash('Аренда не найдена.', 'error')
        return redirect(url_for('rental'))
    try:
        for c in list(r.charges):
            db.session.delete(c)
        db.session.delete(r)
        db.session.commit()
        log_action('rental_delete', 'Rental', rid, 'hard delete with charges')
        flash('Аренда удалена и не учитывается в статистике.', 'success')
    except Exception:
        db.session.rollback()
        flash('Ошибка удаления аренды.', 'error')
    return redirect(url_for('rental'))

@app.route('/rental/item', methods=['GET'])
@login_required
@roles_required('admin','cashier')
def rental_item_pick():
    item_id = request.args.get('item_id')
    if item_id:
        try:
            iid = int(item_id)
        except Exception:
            flash('Неверный идентификатор позиции.', 'error')
            return redirect(url_for('rental'))
        return redirect(url_for('rental_item_detail', item_id=iid))
    flash('Выберите позицию для просмотра статистики.', 'error')
    return redirect(url_for('rental'))

@app.route('/rental/item/<int:item_id>', methods=['GET'])
@login_required
@roles_required('admin','cashier')
def rental_item_detail(item_id: int):
    item = db.session.get(RentalItem, item_id)
    if not item:
        flash('Позиция не найдена.', 'error')
        return redirect(url_for('rental'))
    rentals = Rental.query.filter_by(item_id=item.id).order_by(Rental.started_at.desc()).all()
    total_rentals = len(rentals)
    active = [r for r in rentals if r.returned_at is None]
    closed = [r for r in rentals if r.returned_at is not None]
    total_sum = sum((r.total_price or 0) for r in rentals)
    last_started_at = rentals[0].started_at if rentals else None

    # durations in hours for closed rentals
    import math
    dur_hours = []
    for r in closed:
        try:
            dt = (r.returned_at - r.started_at).total_seconds() / 3600.0
            if dt >= 0:
                dur_hours.append(dt)
        except Exception:
            pass
    total_hours = sum(dur_hours)
    avg_hours = (total_hours / len(dur_hours)) if dur_hours else 0

    # extensions count
    ext_count = 0
    for r in rentals:
        ext_count += sum(1 for c in r.charges if (c.label or '').lower().startswith('extend'))

    # top renters by phone with optional period filter (defaults to current month)
    from collections import defaultdict
    from datetime import datetime as _dt, timedelta as _td, date as _date
    f_str = (request.args.get('from') or '').strip()
    t_str = (request.args.get('to') or '').strip()
    today = _date.today()
    default_from = _date(today.year, today.month, 1)
    df = default_from
    dt = today
    try:
        if f_str:
            df = _dt.strptime(f_str, '%Y-%m-%d').date()
    except Exception:
        df = default_from
    try:
        if t_str:
            dt = _dt.strptime(t_str, '%Y-%m-%d').date()
    except Exception:
        dt = today
    # Inclusive range [df, dt], compare by started_at
    df_dt = _dt.combine(df, _dt.min.time())
    dt_dt_excl = _dt.combine(dt, _dt.min.time()) + _td(days=1)
    rentals_filtered = [r for r in rentals if (r.started_at and df_dt <= r.started_at < dt_dt_excl)]

    renters = defaultdict(lambda: {'count': 0, 'sum': 0.0})
    for r in rentals_filtered:
        key = (r.renter_phone or '-').strip() or '-'
        renters[key]['count'] += 1
        renters[key]['sum'] += (r.total_price or 0)
    top_renters = sorted(renters.items(), key=lambda kv: (-kv[1]['sum'], -kv[1]['count']))[:5]
    period_sum = round(sum(v['sum'] for v in renters.values()), 2)

    # limit history
    history = rentals[:50]

    return render_template(
        'rental_item.html',
        item=item,
        total_rentals=total_rentals,
        active_count=len(active),
        closed_count=len(closed),
        total_sum=round(total_sum, 2),
        last_started_at=last_started_at,
        total_hours=round(total_hours, 2),
        avg_hours=round(avg_hours, 2),
        ext_count=ext_count,
        top_renters=top_renters,
        period_sum=period_sum,
    filter_from=df.strftime('%Y-%m-%d'),
    filter_to=dt.strftime('%Y-%m-%d'),
        history=history,
    )

# ---- POS ----
@app.route('/pos', methods=['GET','POST'])
@login_required
def pos():
    q = request.args.get('q','').strip()
    tab = request.args.get('tab','products')
    if q:
        raw = q
        like = f"%{raw}%"
        norm = _normalize_text(raw)
        tokens = [t for t in raw.lower().split() if t]
        norm_tokens = [t for t in norm.split() if t] or [norm]

        from sqlalchemy import func, case

        # subsequence pattern like %f%4% for "f4"
        def subseq(s): 
            return "%%" + "%%".join(list(s)) + "%%" if s else "%%"

        # base conditions (any token matches across name/sku/barcode)
        conds_per_token = []
        for t, nt in zip(tokens or [raw.lower()], norm_tokens):
            t_like = f"%{t}%"
            nt_like = f"%{nt}%"
            nt_sub = subseq(nt)
            conds_per_token.append(
                or_(
                    Product.name.ilike(t_like),
                    Product.sku.ilike(t_like),
                    Product.barcode.ilike(t_like),
                    Product.search_text.like(nt_like),
                    Product.search_text.like(nt_sub)
                )
            )
        cond = and_(Product.is_archived == False, *conds_per_token)

        # ranking: exact prefix match > substring > fuzzy subseq > others
        rank = case(
            (Product.name.ilike(f"{raw}%"), 0),
            (Product.name.ilike(like), 1),
            (Product.search_text.like(subseq(norm)), 2),
            else_=3
        )
        # stock first (in-stock), then rank, then stock desc, then name
        stock_rank = case((Product.stock_qty > 0, 0), else_=1)
        products = (Product.query
                    .filter(cond)
                    .order_by(stock_rank.asc(), rank.asc(), Product.stock_qty.desc(), Product.name.asc())
                    .limit(100).all())
    else:
        from sqlalchemy import case
        stock_rank = case((Product.stock_qty > 0, 0), else_=1)
        products = (Product.query
                    .filter(Product.is_archived == False)
                    .order_by(stock_rank.asc(), Product.stock_qty.desc(), Product.created_at.desc())
                    .limit(50).all())

    if request.method == 'POST':
        product = Product.query.get_or_404(int(request.form.get('product_id')))
        if product.is_archived:
            flash("Товар в архиве и недоступен для продажи.", "error")
            return redirect(url_for('pos', q=q))
        qty = int(request.form.get('quantity') or 1)
        if qty <= 0:
            flash("Количество должно быть больше нуля.", "error")
            return redirect(url_for('pos', q=q))
        if product.stock_qty < qty:
            flash("Недостаточно товара на складе.", "error")
            return redirect(url_for('pos', q=q))

        unit = product.base_price
        dtype = request.form.get('discount_type') or None
        dval_raw = request.form.get('discount_value')
        dval = float(dval_raw) if dval_raw not in (None, "") else None
        final = unit
        if dtype == 'percent':
            if dval is None or dval < 0 or dval > 100:
                flash("Скидка в % должна быть от 0 до 100.", "error")
                return redirect(url_for('pos', q=q))
            final = unit * (1 - dval/100.0)
        elif dtype == 'fixed':
            if dval is None or dval < 0:
                flash("Скидка в сумме должна быть неотрицательной.", "error")
                return redirect(url_for('pos', q=q))
            final = unit - dval
        if final < 0:
            flash("Итоговая цена не может быть отрицательной.", "error")
            return redirect(url_for('pos', q=q))

        total = round(final * qty, 2)
        profit = round((final - product.cost_price) * qty, 2)

        product.stock_qty -= qty
        sale = Sale(product_id=product.id, quantity=qty, unit_price_sold=unit,
                    discount_type=dtype, discount_value=dval,
                    final_unit_price=final, total_price=total, profit=profit,
                    cashier=current_user.username)
        db.session.add(sale); db.session.commit()
        log_action('sale_create','Sale',sale.id,f'{product.name} x{qty} for {total:.2f}')
        flash(f"Продажа оформлена: {product.name} x{qty}. Сумма: {total:.2f}", "success")
        return redirect(url_for('pos', q=''))

    # Build cart preview
    cart = session.get('cart') or []
    # Enrich cart with product info
    cart_view = []
    subtotal = 0.0
    for item in cart:
        p = db.session.get(Product, item.get('product_id'))
        if not p: 
            continue
        qty = int(item.get('quantity') or 1)
        unit = float(p.base_price)
        dtype = item.get('discount_type') or None
        dval = item.get('discount_value')
        dval = float(dval) if dval not in (None, "") else None
        final = unit
        if dtype == 'percent' and dval is not None:
            final = unit * (1 - max(0.0, min(100.0, dval))/100.0)
        elif dtype == 'fixed' and dval is not None:
            final = unit - max(0.0, dval)
        line_total = round(final * qty, 2)
        subtotal += line_total
        cart_view.append({
            'product': p,
            'quantity': qty,
            'unit': unit,
            'discount_type': dtype,
            'discount_value': dval,
            'final': final,
            'line_total': line_total,
        })
    subtotal = round(subtotal, 2)
    return render_template('pos.html', products=products, q=q, cart=cart_view, subtotal=subtotal, tab=tab)

# ---- POS Return ----
@app.route('/pos/return', methods=['POST'])
@login_required
def pos_return():
    q = request.args.get('q','').strip()
    product = Product.query.get_or_404(int(request.form.get('product_id')))
    if product.is_archived:
        flash("Товар в архиве и недоступен для продажи.", "error")
        return redirect(url_for('pos', q=q))
    try:
        qty = int(request.form.get('quantity') or 0)
    except Exception:
        qty = 0
    if qty <= 0:
        flash("Количество должно быть больше нуля.", "error")
        return redirect(url_for('pos', q=q))
    note = (request.form.get('note') or '').strip()

    unit = product.base_price
    final = unit
    total = -round(final * qty, 2)
    profit = -round((final - product.cost_price) * qty, 2)

    product.stock_qty += qty
    sale = Sale(product_id=product.id, quantity=-qty, unit_price_sold=unit,
                discount_type=None, discount_value=None,
                final_unit_price=final, total_price=total, profit=profit,
                cashier=current_user.username, note=note)
    db.session.add(sale)
    db.session.commit()
    log_action('sale_return','Sale', sale.id, f'{product.name} x{qty}; note={note[:120]}')
    flash(f"Возврат оформлен: {product.name} x{qty}. Сумма: {total:.2f}", "success")
    return redirect(url_for('pos', q=q))

# ---- POS Cart ----
def _get_cart():
    return session.get('cart') or []

def _save_cart(cart):
    session['cart'] = cart

@app.route('/pos/cart/add', methods=['POST'])
@login_required
def pos_cart_add():
    q = request.args.get('q','').strip()
    pid = int(request.form.get('product_id'))
    qty = int(request.form.get('quantity') or 1)
    if qty <= 0:
        flash("Количество должно быть больше нуля.", "error")
        return redirect(url_for('pos', q=q))
    dtype = request.form.get('discount_type') or None
    dval_raw = request.form.get('discount_value')
    dval = float(dval_raw) if dval_raw not in (None, "") else None
    p = Product.query.get_or_404(pid)
    if p.is_archived:
        flash("Товар в архиве и недоступен для продажи.", "error")
        return redirect(url_for('pos', q=q))
    cart = _get_cart()
    # merge by product and discount settings
    merged = False
    for it in cart:
        if it.get('product_id') == pid and (it.get('discount_type') or None) == dtype and (float(it.get('discount_value')) if it.get('discount_value') not in (None, "") else None) == dval:
            it['quantity'] = int(it.get('quantity') or 0) + qty
            merged = True
            break
    if not merged:
        cart.append({'product_id': pid, 'quantity': qty, 'discount_type': dtype, 'discount_value': dval})
    _save_cart(cart)
    flash(f"Добавлено в корзину: {p.name} x{qty}.", "success")
    return redirect(url_for('pos', q=q))

@app.route('/pos/cart/update', methods=['POST'])
@login_required
def pos_cart_update():
    q = request.args.get('q','').strip()
    idx = int(request.form.get('index') or -1)
    qty = int(request.form.get('quantity') or 0)
    cart = _get_cart()
    if 0 <= idx < len(cart):
        if qty <= 0:
            cart.pop(idx)
        else:
            cart[idx]['quantity'] = qty
    _save_cart(cart)
    return redirect(url_for('pos', q=q, tab='cart'))

@app.route('/pos/cart/clear', methods=['POST'])
@login_required
def pos_cart_clear():
    q = request.args.get('q','').strip()
    _save_cart([])
    flash("Корзина очищена.", "success")
    return redirect(url_for('pos', q=q, tab='cart'))

@app.route('/pos/cart/checkout', methods=['POST'])
@login_required
def pos_cart_checkout():
    q = request.args.get('q','').strip()
    cart = _get_cart()
    if not cart:
        flash("Корзина пуста.", "error")
        return redirect(url_for('pos', q=q, tab='cart'))
    # Build products and validate stock
    items = []
    subtotal = 0.0
    for it in cart:
        p = Product.query.get(it.get('product_id'))
        if not p:
            continue
        qty = int(it.get('quantity') or 0)
        if qty <= 0:
            continue
        if p.is_archived:
            flash(f"Товар {p.name} в архиве и недоступен.", "error")
            return redirect(url_for('pos', q=q, tab='cart'))
        unit = float(p.base_price)
        dtype = it.get('discount_type') or None
        dval = it.get('discount_value')
        dval = float(dval) if dval not in (None, "") else None
        final = unit
        if dtype == 'percent' and dval is not None:
            if dval < 0 or dval > 100:
                flash("Скидка в % должна быть от 0 до 100.", "error")
                return redirect(url_for('pos', q=q, tab='cart'))
            final = unit * (1 - dval/100.0)
        elif dtype == 'fixed' and dval is not None:
            if dval < 0:
                flash("Скидка в сумме должна быть неотрицательной.", "error")
                return redirect(url_for('pos', q=q, tab='cart'))
            final = unit - dval
        if final < 0:
            flash("Итоговая цена не может быть отрицательной.", "error")
            return redirect(url_for('pos', q=q, tab='cart'))
        line_total = round(final * qty, 2)
        subtotal += line_total
        items.append((p, qty, unit, dtype, dval, final, line_total))
        if p.stock_qty < qty:
            flash(f"Недостаточно товара на складе: {p.name}.", "error")
            return redirect(url_for('pos', q=q, tab='cart'))
    subtotal = round(subtotal, 2)

    cash_raw = (request.form.get('cash_received') or '').strip()
    try:
        cash_received = float((cash_raw or '0').replace(',', '.'))
    except Exception:
        cash_received = 0.0
    if cash_raw == '':
        cash_received = subtotal
    if cash_received < subtotal:
        flash("Недостаточно средств для оплаты.", "error")
        return redirect(url_for('pos', q=q))
    change = round(cash_received - subtotal, 2)

    # Apply
    for p, qty, unit, dtype, dval, final, line_total in items:
        p.stock_qty -= qty
        sale = Sale(product_id=p.id, quantity=qty, unit_price_sold=unit,
                    discount_type=dtype, discount_value=dval,
                    final_unit_price=final, total_price=line_total,
                    profit=round((final - p.cost_price) * qty, 2),
                    cashier=current_user.username)
        db.session.add(sale)
    db.session.commit()
    _save_cart([])
    log_action('sale_checkout','Sale', None, f'items={len(items)}, total={subtotal:.2f}, cash={cash_received:.2f}, change={change:.2f}')
    flash(f"Продажа оформлена. Итого: {subtotal:.2f}. Получено: {cash_received:.2f}. Сдача: {change:.2f}", "success")
    return redirect(url_for('pos', q='', tab='products'))

# CLI
@app.cli.command("init-db")
def init_db_cmd():
    init_defaults()
    backfill_search_text()
    print("Database initialized successfully.")

# ---- Export Products (Excel) ----
@app.route('/admin/products/export', methods=['GET'])
@login_required
@roles_required('admin')
def admin_products_export():
    import pandas as pd
    from io import BytesIO
    
    # Получаем все товары
    products = Product.query.all()
    
    # Создаем DataFrame
    data = []
    for p in products:
        data.append({
            'ID': p.id,
            'Название': p.name,
            'SKU': p.sku,
            'Штрих-код': p.barcode,
            'Цена закуп.': p.cost_price,
            'Цена прод.': p.base_price,
            'Остаток': p.stock_qty,
            'Архив': 'Да' if p.is_archived else 'Нет',
            'Последнее обновление': p.created_at
        })
    
    df = pd.DataFrame(data)
    
    # Создаем Excel файл в памяти
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Склад')
        worksheet = writer.sheets['Склад']
        # Автоматическая ширина колонок
        for i, col in enumerate(df.columns):
            # Преобразуем все значения в строки для корректного расчета длины
            col_values = df[col].astype(str)
            # Находим максимальную длину значения в колонке
            max_length = max(col_values.str.len().max(), len(col))
            # Устанавливаем ширину колонки с небольшим отступом
            worksheet.set_column(i, i, max_length + 2)
            
        # Добавляем форматирование
        workbook = writer.book
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'align': 'center',
            'bg_color': '#D9E1F2',
            'border': 1
        })
        
        # Применяем форматирование к заголовкам
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            
        # Форматы для цен
        price_format = workbook.add_format({'num_format': '# ##0.00'})
        worksheet.set_column('E:F', None, price_format)  # Для колонок с ценами
        
    output.seek(0)
    
    # Отправляем файл
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
    # Use MSK date for friendly filename
    download_name=f'inventory_{datetime.now(ZoneInfo("Europe/Moscow")).strftime("%Y%m%d")}.xlsx'
    )

# ---- Import Products (Excel/CSV) ----
@app.route('/admin/products/import', methods=['POST'])
@login_required
@roles_required('admin')
def admin_products_import():
    # Валидация файла (ранний выход с понятным сообщением)
    file = request.files.get('file')
    if not file or file.filename == '':
        flash('Файл не выбран.', 'error')
        return redirect(url_for('admin_products'))
    file = request.files.get('file')
    update_if_exists = request.form.get('update_if_exists') == '1'
    if not file or file.filename == '':
        flash('Файл не выбран.', 'error')
        return redirect(url_for('admin_products'))
    ext = os.path.splitext(file.filename)[1].lower()
    allowed_ext = {'.xls', '.xlsx', '.csv', '.xlsm', '.xltx', '.xltm'}
    if ext not in allowed_ext:
        flash('Неверный формат файла. Разрешено: .xls, .xlsx', 'error')
        return redirect(url_for('admin_products'))
    added = 0
    updated = 0
    skipped = 0

    def norm_header(s):
        try:
            s = '' if s is None else str(s)
        except Exception:
            s = ''
        return s.strip().lower()

    def to_float(x):
        try:
            return float(str(x).replace(',', '.'))
        except Exception:
            return None

    def to_int(x):
        try:
            return int(float(str(x).replace(',', '.')))
        except Exception:
            return None

    rows = []
    try:
        if ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
            import openpyxl
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            values_rows = list(ws.iter_rows(values_only=True))
            # автоопределение строки заголовков — самая заполненная строка в первых 50
            header_row = 0
            headers = []
            best = -1
            for ridx, row in enumerate(values_rows[:50]):
                heads_try = [norm_header(v) for v in row]
                score = sum(1 for h in heads_try if h)
                if score > best:
                    best = score
                    header_row = ridx
                    headers = heads_try
            # synonyms
            synonyms = {
                'name': ['name','товар','наименование','название','product'],
                'sku': ['sku','код','артикул','product_code','код товара'],
                'barcode': ['barcode','штрихкод','штрих-код','штрих код','ean','upc'],
                'cost_price': ['cost_price','cost','себестоимость','закуп','закупка','закупочная','покупка','закуп.'],
                'base_price': ['base_price','price','цена','цена продажи','розничная цена','розница','продажа'],
                'stock_qty': ['stock_qty','qty','количество','кол-во','остаток','доступно','остаток на складе','кол-во на складе'],
                'image_url': ['image_url','image','photo','фото','изображение','url']
            }
            # extend synonyms with clean Russian variants
            synonyms_clean = {
                'name': ['name','товар','наименование','название','product'],
                'sku': ['sku','код','артикул','product_code','код товара'],
                'barcode': ['barcode','штрихкод','штрих-код','штрих код','ean','upc'],
                'cost_price': ['cost_price','cost','себестоимость','закуп','закупка','закупочная','покупка','закуп.'],
                'base_price': ['base_price','price','цена','цена продажи','розничная цена','розница','продажа'],
                'stock_qty': ['stock_qty','qty','количество','кол-во','остаток','доступно','остаток на складе','кол-во на складе'],
                'image_url': ['image_url','image','photo','фото','изображение','url']
            }
            for k, alts in synonyms_clean.items():
                synonyms[k] = list({*synonyms.get(k, []), *alts})
            idx = {}
            for k, alts in synonyms.items():
                for a in alts:
                    if a in headers:
                        idx[k] = headers.index(a); break
            # Доп. автосопоставление по русским заголовкам (xlsx)
            try:
                idx_auto = {}
                for i, h in enumerate(headers):
                    if any(x in h for x in ('наимен','товар','назван')): idx_auto['name'] = i
                    if ('код' in h and 'штрих' not in h) or 'артикул' in h: idx_auto.setdefault('sku', i)
                    if 'штрих' in h: idx_auto.setdefault('barcode', i)
                    if any(x in h for x in ('себесто','закуп','покуп')): idx_auto.setdefault('cost_price', i)
                    if ('цена' in h and 'сумм' not in h) or 'розниц' in h or 'продаж' in h: idx_auto.setdefault('base_price', i)
                    if any(x in h for x in ('остаток','доступ','кол-во','количество')): idx_auto.setdefault('stock_qty', i)
                for k,v in idx_auto.items():
                    if k not in idx:
                        idx[k] = v
            except Exception:
                pass
            # Доп. автосопоставление по русским заголовкам (xls)
            try:
                idx_auto = {}
                for i, h in enumerate(headers):
                    if any(x in h for x in ('наимен','товар','назван')): idx_auto['name'] = i
                    if ('код' in h and 'штрих' not in h) or 'артикул' in h: idx_auto.setdefault('sku', i)
                    if 'штрих' in h: idx_auto.setdefault('barcode', i)
                    if any(x in h for x in ('себесто','закуп','покуп')): idx_auto.setdefault('cost_price', i)
                    if ('цена' in h and 'сумм' not in h) or 'розниц' in h or 'продаж' in h: idx_auto.setdefault('base_price', i)
                    if any(x in h for x in ('остаток','доступ','кол-во','количество')): idx_auto.setdefault('stock_qty', i)
                for k,v in idx_auto.items():
                    if k not in idx:
                        idx[k] = v
            except Exception:
                pass

            # Доп. финальный русскоязычный фоллбэк сопоставления
            try:
                idx_ru = {}
                for i, h in enumerate(headers):
                    if any(x in h for x in ('наимен','товар','назван')): idx_ru['name'] = i
                    if ('код' in h and 'штрих' not in h) or 'артикул' in h: idx_ru.setdefault('sku', i)
                    if 'штрих' in h: idx_ru.setdefault('barcode', i)
                    if any(x in h for x in ('себесто','закуп','покуп')): idx_ru.setdefault('cost_price', i)
                    if ('цена' in h and 'сумм' not in h) or 'розниц' in h or 'продаж' in h: idx_ru.setdefault('base_price', i)
                    if any(x in h for x in ('остаток','доступ','кол-во','количество')): idx_ru.setdefault('stock_qty', i)
                for k, v in idx_ru.items():
                    if k not in idx:
                        idx[k] = v
            except Exception:
                pass

            for r in values_rows[header_row+1:]:
                def get(k):
                    i = idx.get(k)
                    return r[i] if i is not None and i < len(r) else None
                name = (get('name') or '').strip() if get('name') else ''
                if not name:
                    # выбрать самое длинное текстовое поле как имя товара
                    cand = None
                    for val in r:
                        if isinstance(val, str) and len(val.strip()) > 3:
                            if cand is None or len(val.strip()) > len(cand):
                                cand = val.strip()
                    if cand:
                        name = cand
                    else:
                        skipped += 1; continue
                sku = (str(get('sku')).strip() if get('sku') is not None else None) or None
                barcode = (str(get('barcode')).strip() if get('barcode') is not None else None) or None
                image_url = (str(get('image_url')).strip() if get('image_url') is not None else None) or None
                cost = to_float(get('cost_price'))
                base = to_float(get('base_price'))
                qty = to_int(get('stock_qty'))
                rows.append(dict(name=name, sku=sku, barcode=barcode, image_url=image_url, cost_price=cost, base_price=base, stock_qty=qty))
        elif ext == '.xls':
            try:
                import xlrd
            except Exception:
                flash('Для импорта .xls установите пакет xlrd==1.2.0', 'error')
                return redirect(url_for('admin_products'))
            content = file.read()
            wb = xlrd.open_workbook(file_contents=content)
            sh = wb.sheet_by_index(0)
            # try to auto-detect header row (look for RU/EN keywords)
            header_row = 0
            headers = [norm_header(sh.cell_value(0, c)) for c in range(sh.ncols)]
            try:
                scan_limit = min(50, sh.nrows)
            except Exception:
                scan_limit = 50
            best_count = -1
            for r in range(scan_limit):
                heads_try = [norm_header(sh.cell_value(r, c)) for c in range(sh.ncols)]
                count = sum(1 for h in heads_try if h)
                if count > best_count:
                    best_count = count
                    header_row = r
                    headers = heads_try
            synonyms = {
                'name': ['name','товар','наименование','название','product'],
                'sku': ['sku','код','артикул','product_code','код товара'],
                'barcode': ['barcode','штрихкод','штрих-код','штрих код','ean','upc'],
                'cost_price': ['cost_price','cost','себестоимость','закуп','закупка','закупочная','покупка','закуп.'],
                'base_price': ['base_price','price','цена','цена продажи','розничная цена','розница','продажа'],
                'stock_qty': ['stock_qty','qty','количество','кол-во','остаток','доступно','остаток на складе','кол-во на складе'],
                'image_url': ['image_url','image','photo','фото','изображение','url']
            }
            idx = {}
            for k, alts in synonyms.items():
                for a in alts:
                    if a in headers:
                        idx[k] = headers.index(a); break
            # Fallback: fuzzy map for common RU headers if exact match failed
            if 'name' not in idx:
                for i, h in enumerate(headers):
                    if any(x in h for x in ('наимен','товар','назван')): idx['name'] = i
                    if ('код' in h and 'штрих' not in h) or 'артикул' in h: idx.setdefault('sku', i)
                    if 'штрих' in h: idx.setdefault('barcode', i)
                    if any(x in h for x in ('себесто', 'закуп', 'покуп')): idx.setdefault('cost_price', i)
                    if ('цена' in h and 'сумм' not in h) or 'розниц' in h or 'продаж' in h: idx.setdefault('base_price', i)
                    if any(x in h for x in ('остаток', 'доступ', 'кол-во', 'количество')): idx.setdefault('stock_qty', i)
            # Доп. финальный русскоязычный фоллбэк сопоставления
            try:
                idx_ru = {}
                for i, h in enumerate(headers):
                    if any(x in h for x in ('наимен','товар','назван')): idx_ru['name'] = i
                    if ('код' in h and 'штрих' not in h) or 'артикул' in h: idx_ru.setdefault('sku', i)
                    if 'штрих' in h: idx_ru.setdefault('barcode', i)
                    if any(x in h for x in ('себесто','закуп','покуп')): idx_ru.setdefault('cost_price', i)
                    if ('цена' in h and 'сумм' not in h) or 'розниц' in h or 'продаж' in h: idx_ru.setdefault('base_price', i)
                    if any(x in h for x in ('остаток','доступ','кол-во','количество')): idx_ru.setdefault('stock_qty', i)
                for k, v in idx_ru.items():
                    if k not in idx:
                        idx[k] = v
            except Exception:
                pass

            for r in range(header_row + 1, sh.nrows):
                def get(k):
                    i = idx.get(k)
                    return sh.cell_value(r, i) if i is not None and i < sh.ncols else None
                name = (str(get('name')).strip() if get('name') is not None else '')
                if not name:
                    cand = None
                    row_vals = [sh.cell_value(r, c) for c in range(sh.ncols)]
                    for val in row_vals:
                        if isinstance(val, str) and len(val.strip()) > 3:
                            if cand is None or len(val.strip()) > len(cand):
                                cand = val.strip()
                    if cand:
                        name = cand
                    else:
                        skipped += 1; continue
                sku = (str(get('sku')).strip() if get('sku') is not None else None) or None
                barcode = (str(get('barcode')).strip() if get('barcode') is not None else None) or None
                image_url = (str(get('image_url')).strip() if get('image_url') is not None else None) or None
                cost = to_float(get('cost_price'))
                base = to_float(get('base_price'))
                qty = to_int(get('stock_qty'))
                rows.append(dict(name=name, sku=sku, barcode=barcode, image_url=image_url, cost_price=cost, base_price=base, stock_qty=qty))
        elif ext == '.csv':
            import csv, io
            content = file.read()
            try:
                s = content.decode('utf-8-sig')
            except Exception:
                s = content.decode('cp1251')
            reader = csv.DictReader(io.StringIO(s))
            def lc(d): return {norm_header(k): v for k,v in d.items()}
            for row in reader:
                rr = lc(row)
                # Normalize CSV headers to canonical keys
                csv_map = {
                    'name': ['name','товар','наименование','название','product'],
                    'sku': ['sku','код','артикул','product_code','код товара'],
                    'barcode': ['barcode','штрихкод','штрих-код','штрих код','ean','upc'],
                    'image_url': ['image_url','image','photo','фото','изображение','url'],
                    'cost_price': ['cost_price','cost','себестоимость','закуп','закупка','закупочная','покупка','закуп.'],
                    'base_price': ['base_price','price','цена','цена продажи','розничная цена','розница','продажа'],
                    'stock_qty': ['stock_qty','qty','количество','кол-во','остаток','доступно','остаток на складе','кол-во на складе']
                }
                for key, alts in csv_map.items():
                    if rr.get(key) not in (None, ''):
                        continue
                    for a in alts:
                        if rr.get(a) not in (None, ''):
                            rr[key] = rr.get(a)
                            break
                name = (rr.get('name') or rr.get('название') or rr.get('наименование') or '').strip()
                if not name:
                    skipped += 1; continue
                sku = (rr.get('sku') or rr.get('артикул') or '').strip() or None
                barcode = (rr.get('barcode') or rr.get('штрихкод') or rr.get('штрих-код') or '').strip() or None
                image_url = (rr.get('image_url') or rr.get('image') or rr.get('фото') or '').strip() or None
                cost = to_float(rr.get('cost_price') or rr.get('себестоимость') or rr.get('закупка'))
                base = to_float(rr.get('base_price') or rr.get('цена') or rr.get('price'))
                qty = to_int(rr.get('stock_qty') or rr.get('остаток') or rr.get('количество') or rr.get('qty'))
                rows.append(dict(name=name, sku=sku, barcode=barcode, image_url=image_url, cost_price=cost, base_price=base, stock_qty=qty))
        else:
            flash('Неподдерживаемый формат файла. Используйте .xlsx, .xls или .csv', 'error')
            return redirect(url_for('admin_products'))
    except Exception as e:
        flash(f'Ошибка чтения файла: {e}', 'error')
        return redirect(url_for('admin_products'))

    # Upsert
    for r in rows:
        name = r['name']
        sku = r['sku']
        barcode = r['barcode']
        image_url = r['image_url']
        cost = r['cost_price'] if r['cost_price'] is not None else 0.0
        base = r['base_price'] if r['base_price'] is not None else 0.0
        qty = r['stock_qty'] if r['stock_qty'] is not None else 0

        found = None
        if sku:
            found = Product.query.filter_by(sku=sku).first()
        if not found and barcode:
            found = Product.query.filter_by(barcode=barcode).first()
        if not found:
            found = Product.query.filter_by(name=name).first()

        if found:
            if update_if_exists:
                changed = []
                if found.sku != sku: changed.append(f'sku: {found.sku}->{sku}'); found.sku = sku
                if found.barcode != barcode: changed.append(f'barcode: {found.barcode}->{barcode}'); found.barcode = barcode
                if found.image_url != image_url: changed.append('image_url'); found.image_url = image_url
                if found.cost_price != cost: changed.append(f'cost_price: {found.cost_price}->{cost}'); found.cost_price = cost
                if found.base_price != base: changed.append(f'base_price: {found.base_price}->{base}'); found.base_price = base
                if qty is not None:
                    # here we treat qty as absolute new stock level if provided
                    if found.stock_qty != qty: changed.append(f'stock_qty: {found.stock_qty}->{qty}'); found.stock_qty = qty
                found.search_text = _normalize_text(' '.join(filter(None, [found.name, found.sku, found.barcode])))
                updated += 1
            else:
                skipped += 1
        else:
            p = Product(name=name, sku=sku, barcode=barcode, image_url=image_url, cost_price=cost, base_price=base, stock_qty=qty)
            p.search_text = _normalize_text(' '.join(filter(None, [p.name, p.sku, p.barcode])))
            db.session.add(p)
            added += 1

    db.session.commit()
    log_action('product_import', 'Product', None, f'added={added}, updated={updated}, skipped={skipped}')
    flash(f'Импорт завершён: добавлено {added}, обновлено {updated}, пропущено {skipped}.', 'success')
    return redirect(url_for('admin_products'))

@app.route('/admin/products/import', methods=['GET'])
@login_required
@roles_required('admin')
def admin_products_import_get():
    # Если пользователь перешел по GET, просто отправим на страницу Товаров
    return redirect(url_for('admin_products'))

@app.route('/admin/products/import/template', methods=['GET'])
@login_required
@roles_required('admin')
def admin_products_import_template():
    # Return CSV template
    from flask import make_response
    csv_data = "name,sku,barcode,cost_price,base_price,stock_qty,image_url\n" \
               "Poco F4 Pro,PF4PRO,1234567890123,200.00,260.00,5,https://example.com/poco.jpg\n"
    resp = make_response(csv_data)
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    resp.headers['Content-Disposition'] = 'attachment; filename=products_template.csv'
    return resp



















@app.route('/admin/products/import/preview', methods=['POST'])
@login_required
@roles_required('admin')
def admin_products_import_preview():
    file = request.files.get('file')
    if not file or file.filename == '':
        flash('Файл не выбран.', 'error')
        return redirect(url_for('admin_products'))
    update_if_exists = request.form.get('update_if_exists') == '1'
    ext = os.path.splitext(file.filename)[1].lower()
    allowed_ext = {'.xls', '.xlsx', '.csv', '.xlsm', '.xltx', '.xltm'}
    if ext not in allowed_ext:
        flash('Неверный формат файла. Разрешено: .xls, .xlsx, .csv', 'error')
        return redirect(url_for('admin_products'))

    tmpdir = _ensure_import_tmp()
    token = uuid4().hex
    tmp_path = os.path.join(tmpdir, f"{token}{ext}")
    try:
        file.stream.seek(0)
        data = file.read()
        with open(tmp_path, 'wb') as f:
            f.write(data)
    except Exception:
        flash('Не удалось сохранить временный файл.', 'error')
        return redirect(url_for('admin_products'))

    rows = []
    try:
        with open(tmp_path, 'rb') as f:
            rows = _parse_product_rows_from_filelike(f, ext)
    except Exception as e:
        flash(str(e), 'error')
        return redirect(url_for('admin_products'))

    to_add, to_update, to_skip = [], [], []
    for r in rows:
        name = (r.get('name') or '').strip()
        if not name:
            to_skip.append((r, 'Пустое имя'))
            continue
        existing = Product.query.filter_by(name=name).first()
        if existing:
            if update_if_exists:
                to_update.append((r, existing))
            else:
                to_skip.append((r, 'Существует; обновление выключено'))
        else:
            to_add.append(r)

    sample = rows[:50]
    return render_template('admin_import_preview.html',
                           token=token, ext=ext,
                           update_if_exists=1 if update_if_exists else 0,
                           total=len(rows),
                           add_count=len(to_add),
                           upd_count=len(to_update),
                           skip_count=len(to_skip),
                           sample=sample)

@app.route('/admin/products/import/apply', methods=['POST'])
@login_required
@roles_required('admin')
def admin_products_import_apply():
    token = (request.form.get('token') or '').strip()
    update_if_exists = request.form.get('update_if_exists') == '1'
    if not token:
        flash('Токен предпросмотра не найден.', 'error')
        return redirect(url_for('admin_products'))
    tmpdir = _ensure_import_tmp()
    allowed_ext = ('.xls', '.xlsx', '.csv', '.xlsm', '.xltx', '.xltm')
    tmp_path = None
    for ext in allowed_ext:
        p = os.path.join(tmpdir, f"{token}{ext}")
        if os.path.exists(p):
            tmp_path = p; break
    if not tmp_path:
        flash('Временный файл не найден. Создайте предпросмотр заново.', 'error')
        return redirect(url_for('admin_products'))
    ext = os.path.splitext(tmp_path)[1].lower()

    added = 0; updated = 0; skipped = 0
    try:
        with open(tmp_path, 'rb') as f:
            rows = _parse_product_rows_from_filelike(f, ext)
        for r in rows:
            name = (r.get('name') or '').strip()
            if not name:
                skipped += 1; continue
            sku = r.get('sku')
            barcode = r.get('barcode')
            image_url = r.get('image_url')
            cost = r.get('cost_price')
            base = r.get('base_price')
            qty = r.get('stock_qty')
            found = Product.query.filter_by(name=name).first()
            if found:
                if update_if_exists:
                    if found.sku != sku: found.sku = sku
                    if found.barcode != barcode: found.barcode = barcode
                    if found.image_url != image_url: found.image_url = image_url
                    if cost is not None and found.cost_price != cost: found.cost_price = cost
                    if base is not None and found.base_price != base: found.base_price = base
                    if qty is not None: found.stock_qty = qty
                    parts = [found.name or '', found.sku or '', found.barcode or '']
                    found.search_text = _normalize_text(' '.join([p for p in parts if p]))
                    updated += 1
                else:
                    skipped += 1
            else:
                p = Product(
                    name=name,
                    sku=sku, barcode=barcode, image_url=image_url,
                    cost_price=cost or 0.0, base_price=base or 0.0,
                    stock_qty=qty or 0, search_text=_normalize_text(' '.join([name, sku or '', barcode or '']))
                )
                db.session.add(p)
                added += 1
        db.session.commit()
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        log_action('products_import_apply', 'Product', None, f'+{added}/~{updated}/skip={skipped}')
        flash(f'Импорт завершён. Добавлено: {added}, Обновлено: {updated}, Пропущено: {skipped}', 'success')
        return redirect(url_for('admin_products'))
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при применении импорта: {e}', 'error')
        return redirect(url_for('admin_products'))
# ---- Finance ----
@app.route('/admin/sales/<int:sid>/delete', methods=['POST'])
@login_required
@roles_required('admin')
def admin_sales_delete(sid):
    s = db.session.get(Sale, sid)
    if not s:
        flash('Продажа не найдена.', 'error')
        return redirect(request.referrer or url_for('admin_finance'))
    try:
        # Restore stock
        if s.product_id:
            p = db.session.get(Product, s.product_id)
            if p:
                p.stock_qty = (p.stock_qty or 0) + (s.quantity or 0)
        db.session.delete(s)
        db.session.commit()
        # Audit with before/after stock and sums
        before = None; after = None
        if s.product_id and p:
            before = (p.stock_qty or 0) - (s.quantity or 0)
            after = p.stock_qty
        log_action('sale_delete', 'Sale', sid, f'prod_id={s.product_id} qty={s.quantity} total={s.total_price} stock_before={before} stock_after={after}')
        flash('Продажа удалена. Остатки возвращены на склад.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Не удалось удалить продажу: {e}', 'error')
    return redirect(request.referrer or url_for('admin_finance'))

@app.route('/admin/finance')
@login_required
@roles_required('admin')
def admin_finance():
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    q = Sale.query
    if date_from:
        try:
            q = q.filter(Sale.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            flash("Некорректная дата 'from' (YYYY-MM-DD).", "error")
    if date_to:
        try:
            dt = datetime.fromisoformat(date_to)
            q = q.filter(Sale.created_at < dt.replace(hour=23, minute=59, second=59))
        except ValueError:
            flash("Некорректная дата 'to' (YYYY-MM-DD).", "error")

    sales = q.order_by(Sale.created_at.desc()).all()
    turnover = sum((s.total_price or 0.0) for s in sales)
    gross_profit = sum((s.profit or 0.0) for s in sales)
    gross_margin = (gross_profit / turnover) if turnover else 0.0

    # Discounts (positive sales only)
    discounts_total = 0.0
    for s in sales:
        if s.quantity and s.quantity > 0:
            try:
                discounts_total += max(0.0, (s.unit_price_sold - s.final_unit_price) * s.quantity)
            except Exception:
                pass

    # COGS from stored profit
    cogs = 0.0
    for s in sales:
        qty = s.quantity or 0
        if qty == 0:
            continue
        try:
            unit_cost = (s.final_unit_price or 0.0) - (s.profit / qty)
            cogs += unit_cost * qty
        except Exception:
            pass

    sales_count = sum(1 for s in sales if (s.quantity or 0) > 0)
    returns_count = sum(1 for s in sales if (s.quantity or 0) < 0)
    avg_ticket = (turnover / sales_count) if sales_count else 0.0

    # Inventory snapshot
    products = Product.query.filter(Product.is_archived == False).all()
    inventory_price = sum((p.base_price or 0.0) * (p.stock_qty or 0) for p in products)
    inventory_cost = sum((p.cost_price or 0.0) * (p.stock_qty or 0) for p in products)
    inventory_units = sum((p.stock_qty or 0) for p in products)
    inventory_products = len(products)

    # Aggregations
    from collections import defaultdict, namedtuple
    by_qty = defaultdict(int)
    by_rev = defaultdict(float)
    for s in sales:
        name = (s.product.name if getattr(s, 'product', None) else f"[удален] #{s.product_id}")
        if (s.quantity or 0) > 0:
            by_qty[name] += s.quantity
            by_rev[name] += (s.total_price or 0.0)
    top_by_qty = sorted(by_qty.items(), key=lambda kv: (-kv[1], kv[0]))[:5]
    top_by_rev = sorted(by_rev.items(), key=lambda kv: (-kv[1], kv[0]))[:5]

    Row = namedtuple('Row', 'cashier revenue profit')
    by_cashier = defaultdict(lambda: {'revenue': 0.0, 'profit': 0.0})
    for s in sales:
        key = s.cashier or '-'
        by_cashier[key]['revenue'] += (s.total_price or 0.0)
        by_cashier[key]['profit'] += (s.profit or 0.0)
    cashier_rows = [Row(cashier=k, revenue=v['revenue'], profit=v['profit']) for k, v in by_cashier.items()]
    cashier_rows.sort(key=lambda r: (-r.revenue, r.cashier))

    return render_template(
        'finance.html',
        date_from=date_from, date_to=date_to,
        sales=sales,
        turnover=round(turnover, 2), gross_profit=round(gross_profit, 2), gross_margin=gross_margin,
        discounts_total=round(discounts_total, 2), cogs=round(cogs, 2),
        sales_count=sales_count, returns_count=returns_count, avg_ticket=round(avg_ticket, 2),
        inventory_price=round(inventory_price, 2), inventory_cost=round(inventory_cost, 2),
        inventory_units=inventory_units, inventory_products=inventory_products,
        top_by_qty=top_by_qty, top_by_rev=top_by_rev,
        cashier_rows=cashier_rows,
    )
# ---- Import local report (.xls) ----
@app.route('/admin/products/import/local', methods=['POST'])
@login_required
@roles_required('admin')
def admin_products_import_local():
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    except Exception:
        base_dir = os.getcwd()
    candidates = []
    try:
        for fn in os.listdir(base_dir):
            if fn.lower().startswith('report-stock-') and fn.lower().endswith('.xls'):
                full = os.path.join(base_dir, fn)
                try:
                    mtime = os.path.getmtime(full)
                except Exception:
                    mtime = 0
                candidates.append((mtime, full))
    except Exception:
        candidates = []
    if not candidates:
        flash('���� �� ���᪥ report-stock-*.xls � ����஢.', 'error')
        return redirect(url_for('admin_products'))
    path = sorted(candidates, key=lambda x: -x[0])[0][1]

    try:
        import xlrd
    except Exception:
        flash('��� ������ .xls ��⠭���� ����� xlrd==1.2.0', 'error')
        return redirect(url_for('admin_products'))
    try:
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_index(0)
    except Exception as e:
        flash(f'�訡�� �⥭�� .xls: {e}', 'error')
        return redirect(url_for('admin_products'))

    def norm_header(s):
        try:
            s = '' if s is None else str(s)
        except Exception:
            s = ''
        return s.strip().lower()

    def to_float(x):
        try:
            return float(str(x).replace(',', '.'))
        except Exception:
            return None

    def to_int(x):
        try:
            return int(float(str(x).replace(',', '.')))
        except Exception:
            return None

    header_row = 0
    headers = [norm_header(sh.cell_value(0, c)) for c in range(sh.ncols)]
    best = -1
    scan_limit = min(50, sh.nrows)
    for r in range(scan_limit):
        heads_try = [norm_header(sh.cell_value(r, c)) for c in range(sh.ncols)]
        score = sum(1 for h in heads_try if h)
        if score > best:
            best = score
            header_row = r
            headers = heads_try

    synonyms = {
        'name': ['name','⮢��','обновлено���','��������','product'],
        'sku': ['sku','���','��⨪�','product_code','��� ⮢��'],
        'barcode': ['barcode','���媮�','����-���','���� ���','ean','upc'],
        'cost_price': ['cost_price','cost','ᥡ��⮨�����','����','���㯪�','���㯮筠�','���㯪�','����.'],
        'base_price': ['base_price','price','業�','業� �த���','஧��筠� 業�','஧���','�த���'],
        'stock_qty': ['stock_qty','qty','������⢮','���-��','���⮪','����㯭�','���⮪ �� ᪫���','���-�� �� ᪫���'],
        'image_url': ['image_url','image','photo','��','����ࠦ����','url']
    }
    idx = {}
    for k, alts in synonyms.items():
        for a in alts:
            if a in headers:
                idx[k] = headers.index(a); break

    rows = []
    for r in range(header_row + 1, sh.nrows):
        def get(k):
            i = idx.get(k)
            return sh.cell_value(r, i) if i is not None and i < sh.ncols else None
        name = (str(get('name')).strip() if get('name') is not None else '')
        if not name:
            continue
        sku = (str(get('sku')).strip() if get('sku') is not None else None) or None
        barcode = (str(get('barcode')).strip() if get('barcode') is not None else None) or None
        image_url = None
        cost = to_float(get('cost_price'))
        base = to_float(get('base_price'))
        qty = to_int(get('stock_qty'))
        rows.append(dict(name=name, sku=sku, barcode=barcode, image_url=image_url, cost_price=cost, base_price=base, stock_qty=qty))

    added = updated = skipped = 0
    for r in rows:
        name = r['name']
        sku = r['sku']
        barcode = r['barcode']
        image_url = r['image_url']
        cost = r['cost_price'] if r['cost_price'] is not None else 0.0
        base = r['base_price'] if r['base_price'] is not None else 0.0
        qty = r['stock_qty'] if r['stock_qty'] is not None else 0

        found = None
        if sku:
            found = Product.query.filter_by(sku=sku).first()
        if not found and barcode:
            found = Product.query.filter_by(barcode=barcode).first()
        if not found:
            found = Product.query.filter_by(name=name).first()

        if found:
            if found.sku != sku: found.sku = sku
            if found.barcode != barcode: found.barcode = barcode
            if found.image_url != image_url: found.image_url = image_url
            if found.cost_price != cost: found.cost_price = cost
            if found.base_price != base: found.base_price = base
            if qty is not None: found.stock_qty = qty
            found.search_text = _normalize_text(' '.join(filter(None, [found.name, found.sku, found.barcode])))
            updated += 1
        else:
            p = Product(name=name, sku=sku, barcode=barcode, image_url=image_url, cost_price=cost, base_price=base, stock_qty=qty)
            p.search_text = _normalize_text(' '.join(filter(None, [p.name, p.sku, p.barcode])))
            db.session.add(p)
            added += 1

    db.session.commit()
    log_action('product_import_local', 'Product', None, f'added={added}, updated={updated}, skipped={skipped}, file={os.path.basename(path)}')
    flash(f'Импорт завершён: добавлено {added}, обновлено {updated}, пропущено {skipped}.', 'success')
    return redirect(url_for('admin_products'))

# ---- Users Management ----
@app.route('/admin/users', methods=['GET', 'POST'])
@login_required
@roles_required('admin')
def admin_users():
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''
        role = (request.form.get('role') or 'cashier').strip() or 'cashier'
        if not username or not password:
            flash('�������� ��易⥫쭮.', 'error')
            return redirect(url_for('admin_users'))
        if User.query.filter_by(username=username).first():
            flash('�⮣���� 業� �� ����� ���� ����⥫쭮�.', 'error')
            return redirect(url_for('admin_users'))
        u = User(username=username, role=role)
        u.set_password(password)
        db.session.add(u); db.session.commit()
        log_action('user_create', 'User', u.id, f'{username}/{role}')
        flash('�室 �믮���� �ᯥ譮.', 'success')
        return redirect(url_for('admin_users'))
    users = User.query.order_by(User.id.asc()).all()
    return render_template('admin_users.html', users=users)

@app.route('/admin/users/<int:uid>/delete', methods=['POST'])
@login_required
@roles_required('admin')
def admin_users_delete(uid):
    u = db.session.get(User, uid)
    if not u:
        flash('�⮣���� 業� �� ����� ���� ����⥫쭮�.', 'error')
        return redirect(url_for('admin_users'))
    try:
        if current_user.id == u.id:
            flash('���� �� ��࠭.', 'error')
            return redirect(url_for('admin_users'))
    except Exception:
        pass
    if u.role == 'admin':
        admins_left = User.query.filter_by(role='admin').count()
        if admins_left <= 1:
            flash('�������筮 ⮢�� �� ᪫���.', 'error')
            return redirect(url_for('admin_users'))
    name = u.username
    db.session.delete(u); db.session.commit()
    log_action('user_delete', 'User', uid, f'deleted {name}')
    flash('����� 㤠��.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:uid>/role', methods=['POST'])
@login_required
@roles_required('admin')
def admin_users_update_role(uid):
    u = db.session.get(User, uid)
    if not u:
        flash('Пользователь не найден.', 'error')
        return redirect(url_for('admin_users'))
    new_role = (request.form.get('role') or '').strip()
    if new_role not in ('admin','cashier'):
        flash('Некорректная роль.', 'error')
        return redirect(url_for('admin_users'))
    if u.role == 'admin' and new_role != 'admin':
        admins_left = User.query.filter_by(role='admin').count()
        if admins_left <= 1:
            flash('Должен остаться хотя бы один администратор.', 'error')
            return redirect(url_for('admin_users'))
    old = u.role
    u.role = new_role
    db.session.commit()
    log_action('user_role', 'User', uid, f'{old}->{new_role}')
    flash('Роль обновлена.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:uid>/password', methods=['POST'])
@login_required
@roles_required('admin')
def admin_users_update_password(uid):
    u = db.session.get(User, uid)
    if not u:
        flash('Пользователь не найден.', 'error')
        return redirect(url_for('admin_users'))
    pwd = request.form.get('password') or ''
    if not pwd:
        flash('Пароль не может быть пустым.', 'error')
        return redirect(url_for('admin_users'))
    u.set_password(pwd)
    db.session.commit()
    log_action('user_password', 'User', uid, 'password reset')
    flash('Пароль обновлён.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/dev')
@login_required
@roles_required('admin')
def admin_dev():
    import sys, flask, sqlalchemy
    db_uri = app.config.get('SQLALCHEMY_DATABASE_URI')
    db_path = None
    if isinstance(db_uri, str) and db_uri.startswith('sqlite:///'):
        db_path = db_uri.replace('sqlite:///','')
        if not os.path.isabs(db_path):
            db_path = os.path.join(os.getcwd(), db_path)
    db_size = None
    if db_path and os.path.exists(db_path):
        try:
            db_size = os.path.getsize(db_path)
        except Exception:
            db_size = None
    counts = {
        'users': User.query.count(),
        'products': Product.query.count(),
        'sales': Sale.query.count(),
        'rental_items': RentalItem.query.count(),
        'rentals': Rental.query.count(),
        'rental_charges': RentalCharge.query.count(),
        'logs': AuditLog.query.count(),
    }
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(20).all()
    info = {
        'python': sys.version.split()[0],
        'flask': getattr(__import__('flask'), '__version__', 'n/a'),
        'sqlalchemy': __import__('sqlalchemy').__version__,
        'version': VERSION,
        'db_uri': db_uri,
        'db_path': db_path,
        'db_size': db_size,
    }
    return render_template('dev.html', info=info, counts=counts, logs=logs)

@app.route('/admin/dev/backfill', methods=['POST'])
@login_required
@roles_required('admin')
def admin_dev_backfill():
    backfill_search_text()
    flash('Поисковые индексы обновлены.', 'success')
    return redirect(url_for('admin_dev'))

@app.route('/admin/dev/init', methods=['POST'])
@login_required
@roles_required('admin')
def admin_dev_init():
    init_defaults()
    flash('Инициализация выполнена.', 'success')
    return redirect(url_for('admin_dev'))

@app.route('/admin/dev/db', methods=['GET'])
@login_required
@roles_required('admin')
def admin_dev_db():
    db_uri = app.config.get('SQLALCHEMY_DATABASE_URI')
    if isinstance(db_uri, str) and db_uri.startswith('sqlite:///'):
        path = db_uri.replace('sqlite:///','')
        if not os.path.isabs(path):
            path = os.path.join(os.getcwd(), path)
        if os.path.exists(path):
            return send_file(path, as_attachment=True)
    flash('Загрузка БД доступна только для SQLite.', 'error')
    return redirect(url_for('admin_dev'))
_init_flag = False
@app.before_request
def _app_bootstrap():
    global _init_flag
    if _init_flag:
        return
    try:
        init_defaults()
        backfill_search_text()
    except Exception:
        pass
    _init_flag = True

if __name__ == '__main__':
    with app.app_context():
        init_defaults()
        backfill_search_text()
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=True)
